
import smertnost_puti
import smertnost_begin_sql
import smertnost_variables
import re
from Pokazateli import cell_navigator
import smertnost_from_journal_to_svod

import openpyxl
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill




# создание функции для многоразового использования кода и сокращения кол-ва текста
def process_merged_cells(top_left_corner, value="", merge_width_cur=10, merge_height_cur=7, alignment="center"):
    # Блок для нахождения нижней правой ячейки по заданным ширине и высоте 
    p = re.compile(r'([A-Z]+)(\d+)')
    m = p.findall(top_left_corner)
    # печать буквенного начала ячейки
    #print(m[0][0])
    begin_cell_letters = m[0][0]
    # печать цифр ячейки
    #print(m[0][1])
    begin_cell_numbers = int(m[0][1])
    # находим индекс, чтобы потом складывать его с шириной
    begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[begin_cell_letters] 
    # складываем с шириной и получаем индекс для конечного столбца, который будем указывать буквой
    end_cell_index_for_letters = begin_cell_index_for_letters + merge_width_cur
    # преобразуем индекс стобца в буквы как в эксель
    end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
    # находим номер строки с учетом высоты
    end_cell_numbers = str(begin_cell_numbers + merge_height_cur)
    bottom_right_corner = end_cell_letters + end_cell_numbers
    range_for_cell = top_left_corner + ":" + bottom_right_corner
    
    smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
    # обрамление границ
    for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
        current_cell_letters = cell_navigator.array_columns[range_of_width]    
        prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
        for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
            smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
    smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
    
    smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal=alignment, vertical='center', wrap_text=True)

    smertnost_from_journal_to_svod.ws[top_left_corner].value = value

    return begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell


top_left_corner = "A19"

#########################
# Вывод шаблона 

number_index_month = 0

current_row = 0

smertnost_variables.tekushee_kolichestvo_vsego = 0

first_time = 1


for smertnost_variables.outer_month in smertnost_puti.months:

    value = smertnost_puti.months_short[number_index_month]
    
    if not first_time:
        #index_for_cell_height = begin_cell_numbers + (number_index_month+1)*5
        top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value, merge_width_cur=10)
    
    first_time = 0
    

    current_row = 0
        
    number_index_month += 1



top_left_corner = "M19"

#########################
# Внесение шаблона

number_index_month = 0

current_row = 0

smertnost_variables.tekushee_kolichestvo_vsego = 0

first_time = 1

value1 = smertnost_variables.pokazatel_short

value2 = smertnost_variables.abs_human

for smertnost_variables.outer_month in smertnost_puti.months:
    
    if not first_time:
        top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value1, merge_width_cur=8,merge_height_cur=3)
    
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value2, merge_width_cur=8,merge_height_cur=3)
    
    first_time = 0      
        
    number_index_month += 1



top_left_corner = "W19"


#########################
# Вывод относит. и абсолют показателей по смертности за предыдущий год 

number_index_month = 0

current_row = 0

smertnost_variables.tekushee_kolichestvo_vsego = 0

first_time = 1


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_all_cause_1000_human_dvoetotsie_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_all_cause_1000_human_dvoetotsie_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1


number_index_month = 0
    

top_left_corner = "AF19"


for smertnost_variables.outer_month in smertnost_puti.months:
    
    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_all_cause_1000_human_dvoetotsie_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)
    
    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_all_cause_1000_human_dvoetotsie_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)
    
    
    number_index_month += 1





top_left_corner = "AO19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.column_name_trud_dvoetotsie_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.column_name_trud_dvoetotsie_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1






top_left_corner = "AX19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.column_name_trud_dvoetotsie_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.column_name_trud_dvoetotsie_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1






top_left_corner = "BG19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_BSK_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_BSK_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1






top_left_corner = "BP19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_BSK_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_BSK_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1





top_left_corner = "BY19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Onco_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Onco_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1




top_left_corner = "CH19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Onco_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Onco_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1






top_left_corner = "CQ19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_DTP_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_DTP_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1






top_left_corner = "CZ19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_DTP_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_DTP_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1





top_left_corner = "DI19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Tuber_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Tuber_100_000_human_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1





top_left_corner = "DR19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Tuber_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Tuber_100_000_human_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1




top_left_corner = "EA19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Child_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Child_year_cur + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1




top_left_corner = "EJ19"


#########################
# 

number_index_month = 0


for smertnost_variables.outer_month in smertnost_puti.months:
    
    # 

    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Child_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.pokazatel_short + '"' + " " + \
                                                ")" ):
            temp=0
    
    
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + " " +  \
                                                '"' + smertnost_variables.title_death_Child_year_prev + '"' + " " + \
                                                smertnost_variables.FROM + \
                                                ' ' + smertnost_variables.table_death_pokazateli + ' ' + \
                                                smertnost_variables.WHERE + " " + \
                                                "(" + \
                                                '"' + smertnost_variables.column_name_pokazatal_month + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_puti.months_short[number_index_month] + '"' + " " + \
                                                smertnost_variables.AND + " " + \
                                                '"' + smertnost_variables.column_name_pokazatel + '"' + " " + \
                                                "=" + \
                                                " " + '"' + smertnost_variables.abs_human + '"' + " " + \
                                                ")" ):
        temp=0

    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
      end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, value=row[0], merge_width_cur=7, merge_height_cur=3)
    
    top_left_corner = begin_cell_letters + str( int(end_cell_numbers)+1)

    number_index_month += 1







