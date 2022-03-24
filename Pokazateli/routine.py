
import smertnost_puti
import smertnost_begin_sql
import smertnost_variables
import re
from Pokazateli import cell_navigator
import smertnost_from_journal_to_svod

import openpyxl
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill



top_left_corner = "A1"

merge_width_cur = 20
merge_height_cur = 13

# ***************** 
# Блок для нахождения нижней правой ячейки по заданным ширине и высоте 
p = re.compile(r'([A-Z]+)(\d+)')
m = p.findall(top_left_corner)
# печать буквенного начала ячейки
#print(m[0][0])
begin_cell_letters = m[0][0]
# печать цифр ячейки
#print(m[0][1])
begin_cell_numbers = int(m[0][1])
# находим индекс, чтобы потом складывать его с шириной
begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[begin_cell_letters] 
# складываем с шириной и получаем индекс для конечного столбца, который будем указывать буквой
end_cell_index_for_letters = begin_cell_index_for_letters + merge_width_cur
# преобразуем индекс стобца в буквы как в эксель
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
# находим номер строки с учетом высоты
end_cell_numbers = str(begin_cell_numbers + merge_height_cur)
bottom_right_corner = end_cell_letters + end_cell_numbers
range_for_cell = top_left_corner + ":" + bottom_right_corner

smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
# *******************


# внесение текста
smertnost_from_journal_to_svod.ws[top_left_corner].value = "Наименование"

# центрирование текста
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center')

# обрамление границ
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border



# создание функции для многоразового использования кода и сокращения кол-ва текста
def process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur):
    # Блок для нахождения нижней правой ячейки по заданным ширине и высоте 
    p = re.compile(r'([A-Z]+)(\d+)')
    m = p.findall(top_left_corner)
    # печать буквенного начала ячейки
    #print(m[0][0])
    begin_cell_letters = m[0][0]
    # печать цифр ячейки
    #print(m[0][1])
    begin_cell_numbers = int(m[0][1])
    # находим индекс, чтобы потом складывать его с шириной
    begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[begin_cell_letters] 
    # складываем с шириной и получаем индекс для конечного столбца, который будем указывать буквой
    end_cell_index_for_letters = begin_cell_index_for_letters + merge_width_cur
    # преобразуем индекс стобца в буквы как в эксель
    end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
    # находим номер строки с учетом высоты
    end_cell_numbers = str(begin_cell_numbers + merge_height_cur)
    bottom_right_corner = end_cell_letters + end_cell_numbers
    range_for_cell = top_left_corner + ":" + bottom_right_corner
    return begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell

#print(begin_cell_letters)

begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)

#print(end_cell_letters)
#print(end_cell_letters +  str(begin_cell_numbers))
merge_width_cur = 16
#merge_height_cur = 13
#print(top_left_corner)



begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
# обрамление границ
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border


smertnost_from_journal_to_svod.ws[top_left_corner].value = smertnost_variables.title_death_all_cause_1000_human

smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)

begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Смертность в трудоспособном возрасте, на 100000 населения"



begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Смертность от болезней системы кровообращения, на 100000 населения"


begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Смертность от новообразований, на 100000 населения"


begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Смертность от ДТП, на 100000 населения"


begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Смертность от туберкулеза, на 100000 населения"


begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Младенческая смертность"


top_left_corner = "A15"

merge_width_cur = 20

# ***************** 
# Блок для нахождения нижней правой ячейки по заданным ширине и высоте 
p = re.compile(r'([A-Z]+)(\d+)')
m = p.findall(top_left_corner)
begin_cell_numbers = int(m[0][1])






merge_height_cur = 3


begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[m[0][0]] - 1
end_cell_index_for_letters = begin_cell_index_for_letters
end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
top_left_corner = end_cell_letters +  str(begin_cell_numbers)
begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
    end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
    current_cell_letters = cell_navigator.array_columns[range_of_width]    
    prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
    for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
        smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
        smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

smertnost_from_journal_to_svod.ws[top_left_corner].value = "Период"





merge_width_cur = 7

for item in range(1,8):
    begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
    end_cell_index_for_letters = begin_cell_index_for_letters
    end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
    top_left_corner = end_cell_letters +  str(begin_cell_numbers)
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
    smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
    for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
        current_cell_letters = cell_navigator.array_columns[range_of_width]    
        prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
        for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
            smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
    smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
    smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    smertnost_from_journal_to_svod.ws[top_left_corner].value = smertnost_puti.year_current
    begin_cell_index_for_letters = cell_navigator.array_columns_letters_to_index[end_cell_letters] 
    end_cell_index_for_letters = begin_cell_index_for_letters
    end_cell_letters = cell_navigator.array_columns[end_cell_index_for_letters]
    top_left_corner = end_cell_letters +  str(begin_cell_numbers)
    begin_cell_letters, begin_cell_numbers, begin_cell_index_for_letters, end_cell_index_for_letters, \
        end_cell_letters, end_cell_numbers, bottom_right_corner, range_for_cell = process_merged_cells(top_left_corner, merge_width_cur, merge_height_cur)
    smertnost_from_journal_to_svod.ws.merge_cells(range_for_cell)
    for range_of_width in range(begin_cell_index_for_letters, end_cell_index_for_letters + 1):
        current_cell_letters = cell_navigator.array_columns[range_of_width]    
        prev_current_cell_letters = cell_navigator.array_columns[range_of_width - 1]    
        for range_of_height in range(int(begin_cell_numbers), int(end_cell_numbers) + 1):        
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height)].border = smertnost_from_journal_to_svod.thin_border
            smertnost_from_journal_to_svod.ws[current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
            smertnost_from_journal_to_svod.ws[prev_current_cell_letters + str(range_of_height + 1)].border = smertnost_from_journal_to_svod.only_top_border
    smertnost_from_journal_to_svod.ws[top_left_corner].border = smertnost_from_journal_to_svod.thin_border
    smertnost_from_journal_to_svod.ws[top_left_corner].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    smertnost_from_journal_to_svod.ws[top_left_corner].value = smertnost_puti.year_prev
    



























