
import smertnost_puti
import pandas as pd
import smertnost_begin_sql
import sqlite3 as sql  # для сортировки, фильтрации и изменению информации
import smertnost_variables

import re

import openpyxl

from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill



temp_fetchall_prev = []

# обрабатываю полик
for row in smertnost_puti.cur2_polik.execute('SELECT * FROM "death_prev_for_polik" WHERE ( "Место смерти" = "смерть другое место дом" OR "Место смерти" = "дом" ) '):
  temp_fetchall_prev.append(row)




#print(temp_fetchall)



df_to_excel2_for_polik_prev = \
  pd.read_sql('SELECT * FROM "death_prev_for_polik" WHERE ( "Место смерти" = "смерть другое место дом" OR "Место смерти" = "дом" ) ', smertnost_puti.conn_polik)

df_to_excel2_for_polik_prev = df_to_excel2_for_polik_prev.drop(df_to_excel2_for_polik_prev.columns[[0]], axis=1)


from_excel_death_prev_for_poliklinika = \
    smertnost_puti.path_death_all_years + str(smertnost_puti.year_prev) + "/" + \
    smertnost_puti.poliklinika_deaths_name_excel

df_to_excel2_for_polik_prev.to_excel(from_excel_death_prev_for_poliklinika)



wb = openpyxl.load_workbook(from_excel_death_prev_for_poliklinika)

ws = wb['Sheet1']

ws.delete_cols(1)

wb.save(filename=from_excel_death_prev_for_poliklinika)

# !!!!!!!!!!!!!!!!!!! за текущий год не ставил так как вроде без этого работает (видимо заплатка получилась)
#smertnost_puti.from_excel_death_prev = from_excel_death_prev_for_poliklinika


smertnost_puti.death_df_prev = smertnost_puti.pd.read_excel(from_excel_death_prev_for_poliklinika)

death_df_prev_poliklinika = smertnost_puti.death_df_prev

#print("excel saved and ")




temp_fetchall_current = []


# обрабатываю полик
for row in smertnost_puti.cur2_polik.execute('SELECT * FROM "death_for_polik" WHERE ( "Место смерти" = "смерть другое место дом" OR "Место смерти" = "дом" ) '):
  temp_fetchall_current.append(row)



df_to_excel2_for_polik_current = \
  pd.read_sql('SELECT * FROM "death_for_polik" WHERE ( "Место смерти" = "смерть другое место дом" OR "Место смерти" = "дом" ) ', smertnost_puti.conn_polik)

df_to_excel2_for_polik_current = df_to_excel2_for_polik_current.drop(df_to_excel2_for_polik_current.columns[[0]], axis=1)

#print(df_to_excel2_for_polik)
#print("1111111111111222222222")
#print(smertnost_puti.from_excel_death_prev)
#print("3333333334444444444")

from_excel_death_current_for_poliklinika = \
    smertnost_puti.path_death_all_years + str(smertnost_puti.year_current) + "/" + \
    smertnost_puti.poliklinika_deaths_name_excel


#temp_polik_path = "/home/user/Documents/Внутренние отчеты/Периодические/Смертность/Смертность по МО/2021/Журнал регистрации поликлинической смерти для программы.xlsx"


df_to_excel2_for_polik_current.to_excel(from_excel_death_current_for_poliklinika)




wb = openpyxl.load_workbook(from_excel_death_current_for_poliklinika)

ws = wb['Sheet1']

ws.delete_cols(1)

wb.save(filename=from_excel_death_current_for_poliklinika)


smertnost_puti.death_df = smertnost_puti.pd.read_excel(from_excel_death_current_for_poliklinika)

death_df_poliklinika = smertnost_puti.death_df




smertnost_puti.cur2_polik.close()




smertnost_puti.death_df.to_sql('death', smertnost_begin_sql.conn, if_exists='replace')
smertnost_puti.death_df_prev.to_sql('death_prev', smertnost_begin_sql.conn, if_exists='replace')