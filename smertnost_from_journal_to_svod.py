
import smertnost_puti
import pandas as pd


if smertnost_puti.poliklinika:
  from Poliklinika import check_polik_death

import smertnost_begin_sql
import sqlite3 as sql  # для сортировки, фильтрации и изменению информации
import smertnost_variables
import re
from Pokazateli import cell_navigator

import template_xls_insert_part1


# обрабатываю полик
#for row in smertnost_puti.cur2_polik.execute('SELECT * FROM "death_prev_for_polik" WHERE ( "Место смерти" = "смерть другое место дом" OR "Место смерти" = "дом" ) '):
#  temp = row[0]


for smertnost_variables.outer_month in smertnost_puti.months:

    # по строкам "всего" предыдущего года
    # ((( переменные для запросов не стал создавать динамически так как это не безопасно )))
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + \
                                            ' ' + smertnost_variables.count_row + ' ' + \
                                            smertnost_variables.FROM + \
                                            ' ' + smertnost_variables.table_death_prev + ' ' + \
                                            smertnost_variables.WHERE + \
                                            ' ' + '"' + smertnost_variables.column_period_po_zags + '"' + ' ' + \
                                            '==' + \
                                            ' ' + '"' + smertnost_variables.outer_month + '"'):
        # учитываем по нарастающей
        smertnost_variables.tekushee_kolichestvo_vsego_prev += row[0]
        # внесение сведений за предыдущий год в свод "умерло всего"
        smertnost_begin_sql.cur2.execute(smertnost_variables.UPDATE + \
                                     " " + smertnost_variables.table_death_svod + " " + \
                                     smertnost_variables.SET + \
                                     " " + smertnost_variables.outer_month + " " + \
                                     "=" + \
                                     " " + "'" + str(smertnost_variables.tekushee_kolichestvo_vsego_prev) + "'" + " " + \
                                     smertnost_variables.WHERE + \
                                     " " + smertnost_variables.column_pokazalel + " " + \
                                     "=" + \
                                     " " + "'" + smertnost_variables.total + " " + str(
        smertnost_puti.year_prev) + " " + smertnost_variables.year_point + "'")


    # по строкам "всего" текущего года
    
    for row in smertnost_begin_sql.cur2.execute(smertnost_variables.SELECT + \
                                            ' ' + smertnost_variables.count_row + ' ' + \
                                            smertnost_variables.FROM + \
                                            ' ' + smertnost_variables.table_death_cur + ' ' + \
                                            smertnost_variables.WHERE + \
                                            ' ' + '"' + smertnost_variables.column_period_po_zags + '"' + ' ' + \
                                            '==' + \
                                            ' ' + '"' + smertnost_variables.outer_month + '"'):
        # учитываем по нарастающей
        smertnost_variables.tekushee_kolichestvo_vsego += row[0]
        # обновляем подстроки если поступили сведения из ЗАГС и
        # переменная smertnost_variables.for_update_only_current_month тогда не равна нулю
        if row[0]:
            smertnost_variables.for_update_only_current_month = smertnost_variables.tekushee_kolichestvo_vsego
            temp_for_sql = smertnost_variables.tekushee_kolichestvo_vsego
        else:
            # в любом случае изменяем значения temp_for_sql из предыдущих строк свода
            temp_for_sql = 0
            # до конца года будут нули по оставшимся месяцам текущего года
            smertnost_variables.for_update_only_current_month = 0
        
        # внесение сведений за текущий год в свод "умерло всего"
        smertnost_begin_sql.cur2.execute(smertnost_variables.UPDATE + " " + smertnost_variables.table_death_svod + " " + \
                                     smertnost_variables.SET + \
                                     " " + smertnost_variables.outer_month + " " + \
                                     "=" + \
                                     " " + "'" + str(temp_for_sql) + "'" + " " + \
                                     smertnost_variables.WHERE + \
                                     " " + smertnost_variables.column_pokazalel + " " + \
                                     "=" + \
                                     " " + "'" + smertnost_variables.total + " " + str(
            smertnost_puti.year_current) + " " + smertnost_variables.year_point + "'")


for smertnost_variables.outer_month in smertnost_puti.months:
    import BSK
    import OIM
    import ONMK
    import IBS
    import GB
    import TS_A
    import P_ONMK
    import Kapilary
    import ONKO
    import ONKO_localiz
    import Tuberculosis
    import VICH
    import Respiratory
    import Digestive
    import Endocrine
    import Senility
    import Mental_disorders
    import Nervous_system
    import Musculoskeletal_system
    import Genitourinary_system
    import Outer_causes

import Total_death


import template_xls_insert_part2


#import Main_percent


from Outercauses import intro

from Outercauses import mkbS

from Outercauses import mkbT


smertnost_begin_sql.conn.commit()

from Pokazateli import intro

from Pokazateli import pokazateli_bsk

from Pokazateli import pokazateli_onco

from Pokazateli import pokazateli_dtp

from Pokazateli import pokazateli_tuberculosis



smertnost_begin_sql.conn.commit()


df_to_excel2 = pd.read_sql('SELECT * FROM death_svod', smertnost_begin_sql.conn)

df_death_pokazateli = pd.read_sql('SELECT * FROM death_pokazateli', smertnost_begin_sql.conn)

df_to_excel2.to_excel(smertnost_puti.result_excel_dist)

# создаем временный файл для хранения информации из журнала смертности
conn = sql.connect(smertnost_begin_sql.temp_sql_db)

import openpyxl

from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill


#currentCell = ws.cell('A1') #or currentCell = ws['A1']
#currentCell.alignment = Alignment(horizontal='center')


wb = openpyxl.load_workbook(smertnost_puti.result_excel_dist)

prov = ['По нозолог. и внеш. прич.']
#prov = ['По нозолог. и внеш. прич.', 'Относит. показ-ли']
#wb = load_workbook('bunny.xlsx') 
ws = wb['Sheet1']
for city in prov:
    wb.copy_worksheet(ws)
    ws = wb.worksheets[-1]
    ws.title = city

if 'Sheet1' in wb.sheetnames:
    wb.remove(wb['Sheet1'])

ws=wb.get_sheet_by_name("По нозолог. и внеш. прич.")

#wb.save('bunny.xlsx')


#wb = openpyxl.Workbook()

#ws = wb.active
######################################################### ws = wb["По нозолог. и внеш. прич."]

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.scale = 88
ws.page_margins.left=0.3
ws.page_margins.right=0.3
ws.page_margins.top=0.3
ws.page_margins.bottom=0.3
#, right=0.4, top=0.612, bottom=0.1, header=0.2, footer=0.3)


#ws.page_setup.fitToHeight = 1
#ws.page_setup.fitToWidth = 1

#dim_holder = DimensionHolder(worksheet=ws)

dims = {}

r = 1

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

only_top_border = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style='thin'), 
                     bottom=Side(style=None))

side = openpyxl.styles.Side(border_style=None)
"""
no_border = openpyxl.styles.borders.Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
)
"""
no_border = openpyxl.styles.borders.Border(
    left=None, 
    right=None, 
    top=None, 
    bottom=None,
)

ws.delete_cols(1)

ws.merge_cells('A1:M1')
ws.merge_cells('A2:M2')
ws.merge_cells('A3:M3')

row_to_merge1 = 0


#sheet['A1'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type = "solid")

number_row = 1

#openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheetObject, paper_size = <someInt>, orientation='landscape')

for row in ws.rows:

    odd_even_cell_for_month = 1

    if number_row == 2 or number_row == 3:
        ws.row_dimensions[number_row].height = 12
    else:
        ws.row_dimensions[number_row].height = 20
    
    if number_row == 159:
        ws.row_dimensions[number_row].height = 40
    #print(row[0].value) 
    
    for cell in row:  
        
        if hasattr(cell, 'column_letter'):
            column_letter = str(cell.column_letter) 
            ws[column_letter + str(number_row)].font = Font(size = 10.5 , name = 'Times New Roman')
             
        #print(number_row)
        if odd_even_cell_for_month % 2 == 1 and odd_even_cell_for_month > 1 and odd_even_cell_for_month < 13:
            if number_row > 5:
                if odd_even_cell_for_month > 2:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type = "solid")

        if number_row % 2 == 1:
            if number_row > 5:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type = "solid")
            

        if number_row < 4 :
            if number_row == 1 :
                ws['A1'].value = "Анализ показателей причин"
                if smertnost_puti.poliklinika:
                    ws['A1'].value += " поликлинической"
                ws['A1'].value += " смертности"
        elif cell.column_letter == "A" :
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else :
            cell.alignment = Alignment(horizontal='center', vertical='center')

        
        """
        #if number_row == 4 or number_row == 31 or number_row == 57 or number_row == 77 or number_row == 108 or number_row == 130 :
        if cell.value == "Показатель" or cell.value == "Январь" or cell.value == "Февраль" or cell.value == "Март" \
                    or cell.value == "Апрель" or cell.value == "Май" or cell.value == "Июнь" or cell.value == "Июль" \
                    or cell.value == "Август" or cell.value == "Сентябрь" or cell.value == "Октябрь" \
                    or cell.value == "Ноябрь" or cell.value == "Декабрь":
            cell.font = Font(size = 12 , name = 'Times New Roman', bold = True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else :
            cell.font = Font(size = 10.5 , name = 'Times New Roman')
        """
        if number_row == 1:
            cell.border = no_border

        elif number_row > 3:
            cell.border = thin_border

        odd_even_cell_for_month += 1        

        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 3))
    
    if row[0].value == "Показатель":
        for column_in_one_row in range(0,13):
            row[column_in_one_row].font = Font(size = 12 , name = 'Times New Roman', bold = True)
            row[column_in_one_row].alignment = Alignment(horizontal='center', vertical='center') 
    elif row[0].value == "*Анализ показателей проводить по нарастающей":
        row_to_merge1 = number_row
        # Ниже под строкой "*Анализ показателей проводить по нарастающей" будет также белый фон
        inter_table_main_and_outer_causes = number_row + 1
    else:
        for column_in_one_row in range(0,13):
            if column_in_one_row == 3 or column_in_one_row == 6 or column_in_one_row == 9 or column_in_one_row == 12:
                row[column_in_one_row].font = Font(size = 12 , name = 'Times New Roman', bold = True) 

    number_row += 1
        
ws['A1'].font = Font(size = 12 , name = 'Times New Roman')

#print(row_to_merge1)
#print(inter_table_main_and_outer_causes)

ws["A" + str(row_to_merge1)].fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type = "solid")

ws["A" + str(inter_table_main_and_outer_causes)].fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type = "solid")

ws.merge_cells("A" + str(row_to_merge1) + ":" + "M" + str(row_to_merge1))


ws.merge_cells("A" + str(row_to_merge1 + 1) + ":" + "M" + str(row_to_merge1 + 2))

if hasattr( ws[ "A" + str(row_to_merge1 + 1) ] , 'border'):
    for x in [chr(i) for i in range(ord('A'),ord('M') + 1 )]:
        ws[x + str(row_to_merge1 + 1) ].border = no_border
        ws[x + str(row_to_merge1 + 2) ].border = no_border


#for x in [chr(i) for i in range(ord('A'),ord('M') + 1 )]:
#    print(x)

"""
        if hasattr(cell, 'column_letter'):
            column_letter = str(cell.column_letter) 
            ws[column_letter + str(number_row)].font = Font(size = 10.5 , name = 'Times New Roman')

ws.row_dimensions[1].border = no_border
ws.row_dimensions[2].border = no_border
ws.row_dimensions[3].border = no_border    
"""     

for col, value in dims.items():
    if col != "A":
        ws.column_dimensions[col].width = value
    else:
        ws.column_dimensions[col].width = 40
"""
number_column = 1

for row in ws.rows:
    #row[0].width = 50
    ws.column_dimensions[number_column].width = 70
    number_column += 1



smertnost_begin_sql.cur2.execute (INSERT_to_death_svod +
                                 "Анализ показателей причин смертности " + \
                                     smertnost_puti.title_organisation + "'" + ", " + smertnost_puti.tire_12 + ")")
"""




prov = ['Относит. показ-ли']
#wb = load_workbook('bunny.xlsx') 

for city in prov:
    wb.copy_worksheet(ws)
    ws = wb.worksheets[-1]
    ws.title = city

ws=wb.get_sheet_by_name("Относит. показ-ли")

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.scale = 88
ws.page_margins.left=0.3
ws.page_margins.right=0.3
ws.page_margins.top=0.3
ws.page_margins.bottom=0.3

ws.delete_cols(1, 100)
ws.delete_rows(1, 1000)


number_row = 1

for x in [chr(i) for i in range(ord('A'),ord('Z') + 1 )]:
    ws.column_dimensions[x].width = 1
    for y in [chr(ii) for ii in range(ord('A'),ord('Z') + 1 )]:
        ws.column_dimensions[x + y].width = 1
#for row in range(1, 10000):
#    ws.column_dimensions[x].width = 1


for item in range(1, 10000):
    ws.row_dimensions[item].height = 5

#ws.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=13)

# это ошибка библиотеки видимо - требуется повторно обьединить чтобы снять баг
ws.merge_cells('A1:M1')
ws.merge_cells('A2:M2')
ws.merge_cells('A3:M3')

ws.unmerge_cells('A1:M1')
ws.unmerge_cells('A2:M2')
ws.unmerge_cells('A3:M3')



#print(df_death_pokazateli["Наименование: месяц"][0])
# >>> print(chr(65))
# A

# в миллиметрах 
#import re
# пример



from Pokazateli import routine

from Pokazateli import sql_table

#print(bottom_right_corner)

"""
#print(end_cell_numbers)
# chr and ord не буду использовать, так как в юникоде не упорядочиваются
# буквы после слияния по возрастанию
begin_cell = m[0][0] +       \
        str( m[0][1] )
print(begin_cell)
print(ord(m[0][0]))
print( int( ord(m[0][0]) + merge_width_cur ) )
end_cell_letters = chr(  int( ord(m[0][0]) +        \
        merge_width_cur )  )
print(end_cell_letters)
"""
"""
ws.merge_cells( begin_cell +  ":" +  end_cell_letters  +       \
        str( (int( m[0][1] +         \
        merge_height_cur) )  )   )
"""




#ws['AC3'].border = thin_border


ws=wb.get_sheet_by_name("По нозолог. и внеш. прич.")



wb.save(filename=smertnost_puti.result_excel_dist)

conn.close()











